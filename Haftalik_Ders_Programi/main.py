import pyodbc
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def get_connection(database=None):
    connection_string = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=DESKTOP-5QE6TBQ\SQLEXPRESS;'
        + (f'DATABASE={database};' if database else '') +
        'Trusted_Connection=yes;'
    )
    conn = pyodbc.connect(connection_string)
    conn.autocommit = True  
    return conn


def database_exists():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sys.databases WHERE name = 'DersProgramiDB'")
    exists = cursor.fetchone()
    conn.close()
    return exists is not None


def create_database():
    if database_exists():
        print("DersProgramiDB zaten mevcut.")
        return

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("CREATE DATABASE DersProgramiDB")
    conn.commit()
    conn.close()
    print("DersProgramiDB oluşturuldu.")


def table_exists(table_name, conn):
    cursor = conn.cursor()
    cursor.execute(f"""
        SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES 
        WHERE TABLE_NAME = '{table_name}'
    """)
    return cursor.fetchone()[0] > 0


def create_tables():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    tables = {
        "Fakulte": """
        CREATE TABLE Fakulte (
            id INT IDENTITY(1,1) PRIMARY KEY,
            fakulte_adi NVARCHAR(255) NOT NULL UNIQUE
        )
        """,
        "Bolumler": """
        CREATE TABLE Bolumler (
            id INT IDENTITY(1,1) PRIMARY KEY,
            fakulte_id INT FOREIGN KEY REFERENCES Fakulte(id),
            fakulte_adi NVARCHAR(255) NOT NULL FOREIGN KEY REFERENCES Fakulte(fakulte_adi),
            bolum_adi NVARCHAR(255) NOT NULL UNIQUE
        )
        """,
        "OgretimGorevlileri": """
        CREATE TABLE OgretimGorevlileri (
            id INT IDENTITY(1,1) PRIMARY KEY,
            fakulte_id INT FOREIGN KEY REFERENCES Fakulte(id),
            ogretim_gorevlisi NVARCHAR(255) NOT NULL UNIQUE,
            pazartesi NVARCHAR(255),
            sali NVARCHAR(255),
            carsamba NVARCHAR(255),
            persembe NVARCHAR(255),
            cuma NVARCHAR(255)
        )
        """,
        "Ogrenciler": """
        CREATE TABLE Ogrenciler (
            id INT IDENTITY(1,1) PRIMARY KEY,
            fakulte_id INT FOREIGN KEY REFERENCES Fakulte(id),
            bolum_id INT FOREIGN KEY REFERENCES Bolumler(id),
            bolum_adi NVARCHAR(255) NOT NULL FOREIGN KEY REFERENCES Bolumler(bolum_adi),
            sinif INT NOT NULL,
            numara NVARCHAR(50) UNIQUE NOT NULL
        )
        """,
        "Dersler": """
        CREATE TABLE Dersler (
            id INT IDENTITY(1,1) PRIMARY KEY,
            fakulte_id INT,
            bolum_id INT,
            sinif INT NOT NULL,
            ders_kodu NVARCHAR(50) NOT NULL,
            ders_adi NVARCHAR(255) NOT NULL,
            ogretim_uyesi_id INT,
            haftalik_saat INT NOT NULL,
            online NVARCHAR(255),
            zorunlu_saat NVARCHAR(255),
            statu NVARCHAR(10) CHECK (statu IN ('LAB', 'NORMAL')),
            FOREIGN KEY (fakulte_id) REFERENCES Fakulte(id) ON DELETE SET NULL,
            FOREIGN KEY (bolum_id) REFERENCES Bolumler(id) ON DELETE SET NULL,
            FOREIGN KEY (ogretim_uyesi_id) REFERENCES OgretimGorevlileri(id) ON DELETE SET NULL
        )
        """,
        "OgrenciDers": """
        CREATE TABLE OgrenciDers (
            id INT IDENTITY(1,1) PRIMARY KEY,
            ogrenci_id INT,
            ogrenci_num NVARCHAR(50),
            ders_id INT,
            FOREIGN KEY (ogrenci_id) REFERENCES Ogrenciler(id) ON DELETE CASCADE,
            FOREIGN KEY (ogrenci_num) REFERENCES Ogrenciler(numara) ON DELETE NO ACTION,
            FOREIGN KEY (ders_id) REFERENCES Dersler(id) ON DELETE CASCADE
        )
        """,
        "Derslikler": """
        CREATE TABLE Derslikler (
            id INT IDENTITY(1,1) PRIMARY KEY,
            derslik_id NVARCHAR(50) UNIQUE NOT NULL,
            kapasite INT NOT NULL,
            statu NVARCHAR(10) CHECK (statu IN ('LAB', 'NORMAL'))
        )
        """
    }

    for table_name, table_query in tables.items():
        if table_exists(table_name, conn):
            print(f"{table_name} tablosu zaten mevcut.")
        else:
            cursor.execute(table_query)

    conn.commit()
    conn.close()
    print("Tablo oluşturma işlemi tamamlandı.")

create_database()
create_tables()


students_file = 'Ogrenciler.xlsx'
# Fakülteler tablosuna veri eklemek için fonksiyon
def insert_faculties_from_excel(students_file):
    df = pd.read_excel(students_file)
    faculties = df['Fakülte'].unique()

    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    for faculty in faculties:
        cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", faculty)
        if cursor.fetchone() is None:
            cursor.execute("INSERT INTO Fakulte (fakulte_adi) VALUES (?)", faculty)
        else:
            print(f"{faculty} fakültesi zaten mevcut.")

    conn.commit()
    conn.close()
    print("Fakülte ekleme işlemleri tamamlandı.")


# Bölümler tablosuna veri eklemek için fonksiyon
def insert_departments_from_excel(students_file):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    df = pd.read_excel(students_file)
    departments = df[['Fakülte', 'Bölüm']].drop_duplicates()

    for index, row in departments.iterrows():
        faculty = row['Fakülte']
        department = row['Bölüm']

        # Fakülte ID'sini al
        cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", faculty)
        faculty_id = cursor.fetchone()

        if faculty_id:
            faculty_id = faculty_id[0]
            cursor.execute("SELECT id FROM Bolumler WHERE bolum_adi = ? AND fakulte_id = ?", department, faculty_id)
            if cursor.fetchone() is None:
                cursor.execute("INSERT INTO Bolumler (fakulte_id, fakulte_adi, bolum_adi) VALUES (?, ?, ?)",
                               faculty_id, faculty, department)
            else:
                print(f"{department} bölümü zaten mevcut.")
        else:
            print(f"{faculty} fakültesi bulunamadı.")

    conn.commit()
    conn.close()
    print("Fakülte ekleme işlemleri tamamlandı.")


# Öğrencileri Excel'den okuyup veritabanına ekleyen fonksiyon
def insert_students_from_excel(students_file):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    df = pd.read_excel(students_file)

    for index, row in df.iterrows():
        faculty_name = row['Fakülte']
        department_name = row['Bölüm']
        student_class = row['Sınıf']
        student_number = row['Numara']

        # Fakülte ID'sini al
        cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", faculty_name)
        faculty_id = cursor.fetchone()

        # Bölüm ID'sini al
        cursor.execute("SELECT id FROM Bolumler WHERE bolum_adi = ? AND fakulte_id = ?", department_name,
                       faculty_id[0] if faculty_id else None)
        department_id = cursor.fetchone()

        if faculty_id and department_id:
            faculty_id = faculty_id[0]
            department_id = department_id[0]

            # Öğrenci zaten var mı kontrol et
            cursor.execute("SELECT id FROM Ogrenciler WHERE numara = ?", student_number)
            if cursor.fetchone() is None:
                cursor.execute("""
                    INSERT INTO Ogrenciler (fakulte_id, bolum_id, bolum_adi, sinif, numara) 
                    VALUES (?, ?, ?, ?, ?)
                """, faculty_id, department_id, department_name, student_class, student_number)
            else:
                print(f"{student_number} numaralı öğrenci zaten mevcut.")
        else:
            print(f"{faculty_name} fakültesi veya {department_name} bölümü bulunamadı.")
    conn.commit()
    conn.close()
    print("Öğrenci ekleme işlemleri tamamlandı.")

insert_faculties_from_excel(students_file)  # Fakülteleri eklemek için
insert_departments_from_excel(students_file)  # Bölümleri eklemek için
insert_students_from_excel(students_file)  # Öğrencileri eklemek için


faculty_members_file = 'OgretimUyeleri.xlsx'
# Öğretim üyelerini Excel'den okuyup veritabanına ekleyen fonksiyon
def insert_faculty_members_from_excel(faculty_members_file):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    df = pd.read_excel(faculty_members_file)
    df.columns = df.columns.str.strip()

    for column in ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']:
        df[column] = df[column].fillna('')

        # Fakülteleri veritabanına ekle
    for _, row in df.iterrows():
        faculty_name = row['Fakülte']
        cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (faculty_name,))
        faculty_id = cursor.fetchone()

        # Eğer fakülte veritabanında yoksa ekle
        if faculty_id is None:
            cursor.execute("INSERT INTO Fakulte (fakulte_adi) VALUES (?)", (faculty_name,))
            faculty_id = cursor.lastrowid
            print(f"{faculty_name} fakültesi veritabanına eklendi.")
        else:
            faculty_id = faculty_id[0]

        teacher_name = row['Öğretim Görevlisi']

        # Aynı öğretim görevlisinin zaten veritabanında olup olmadığını kontrol et
        cursor.execute("""
            SELECT COUNT(*) FROM OgretimGorevlileri 
            WHERE fakulte_id = ? AND ogretim_gorevlisi = ?
        """, (faculty_id, teacher_name))
        existing_record_count = cursor.fetchone()[0]

        if existing_record_count == 0:
            monday = str(row['Pazartesi'])
            tuesday = str(row['Salı'])
            wednesday = str(row['Çarşamba'])
            thursday = str(row['Perşembe'])
            friday = str(row['Cuma'])

            cursor.execute("""
                INSERT INTO OgretimGorevlileri (fakulte_id, ogretim_gorevlisi, pazartesi, sali, carsamba, persembe, cuma)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (faculty_id, teacher_name, monday, tuesday, wednesday, thursday, friday))

        else:
            print(f"{teacher_name} öğretim görevlisi zaten mevcut.")

    conn.commit()
    conn.close()
    print("Öğretim görevlileri başarıyla eklendi.")

insert_faculty_members_from_excel(faculty_members_file)  # Öğretim üyelerini eklemek için


classroom_file = 'Derslikler.xlsx'
# Derslikleri Excel'den okuyup veritabanına ekleyen fonksiyon
def insert_classrooms_from_excel(clasroom_file):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    df = pd.read_excel(clasroom_file)
    df.columns = df.columns.str.strip()

    # Derslikleri veritabanına ekle
    for _, row in df.iterrows():
        classroom_id = row['Derslik_ID']
        capacity = row['Kapasite']
        status = row['Statü']

        # Aynı derslik ID'sinin zaten veritabanında olup olmadığını kontrol et
        cursor.execute("SELECT COUNT(*) FROM Derslikler WHERE derslik_id = ?", (classroom_id,))
        existing_record_count = cursor.fetchone()[0]

        # Eğer mevcut değilse ekle
        if existing_record_count == 0:
            cursor.execute("""
                INSERT INTO Derslikler (derslik_id, kapasite, statu)
                VALUES (?, ?, ?)
            """, (classroom_id, capacity, status))
        else:
            print(f"{classroom_id} dersliği zaten mevcut.")

    conn.commit()
    conn.close()
    print("Derslik ekleme işlemleri tamamlandı.")

insert_classrooms_from_excel(classroom_file)  # Derslikleri eklemek için


courses_file = "Dersler.xlsx"
def insert_courses_from_excel(courses_file):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    df = pd.read_excel(courses_file)

    # Online ve Statü sütunlarını kontrol et eğer boşsa varsayılan değerler ata
    if 'Online' in df.columns:
        df['Online'].fillna('Hayır', inplace=True)
    else:
        df['Online'] = 'Hayır'

    if 'Statü' in df.columns:
        df['Statü'].fillna('NORMAL', inplace=True)
        # Eğer statü LAB ya da NORMAL değilse, 'NORMAL' olarak değiştir
        df['Statü'] = df['Statü'].apply(lambda x: x if x in ['LAB', 'NORMAL'] else 'NORMAL')
    else:
        df['Statü'] = 'NORMAL'

    df['Haftalık Saat'] = pd.to_numeric(df['Haftalık Saat'], errors='coerce').fillna(0).astype(int)
    df['Zorunlu Saat'] = df['Zorunlu Saat'].fillna('0').astype(str)
    for index, row in df.iterrows():
        # Fakülte ID'sini bul
        cursor.execute("SELECT id FROM fakulte WHERE fakulte_adi = ?", (row['Fakülte'],))
        fakulte_id = cursor.fetchone()
        fakulte_id = fakulte_id[0] if fakulte_id else None

        # Bölüm ID'sini bul
        cursor.execute("SELECT id FROM bolumler WHERE bolum_adi = ?", (row['Bölüm'],))
        bolum_id = cursor.fetchone()
        bolum_id = bolum_id[0] if bolum_id else None

        # Öğretim Üyesi ID ve adı al
        cursor.execute("SELECT id FROM ogretimgorevlileri WHERE ogretim_gorevlisi = ?", (row['Öğretim Üyesi'],))
        ogretim_uyesi = cursor.fetchone()

        # Eğer öğretim üyesi yoksa yeni öğretim üyesini ekle
        if ogretim_uyesi is None:
            cursor.execute("""
                INSERT INTO ogretimgorevlileri (ogretim_gorevlisi) 
                VALUES (?)
            """, (row['Öğretim Üyesi'],))
            conn.commit()
            cursor.execute("SELECT id FROM ogretimgorevlileri WHERE ogretim_gorevlisi = ?", (row['Öğretim Üyesi'],))
            ogretim_uyesi = cursor.fetchone()

        ogretim_uyesi_id = ogretim_uyesi[0]
        online = True if str(row['Online']).strip().lower() == "evet" else False

        # Dersin var olup olmadığını kontrol et
        cursor.execute("""
            SELECT id FROM Dersler 
            WHERE bolum_id = ? AND sinif = ? AND ders_kodu = ? AND ders_adi = ?
        """, (bolum_id, row['Sınıf'], row['Ders Kodu'], row['Ders Adı']))
        existing_course = cursor.fetchone()

        if existing_course:
            print(f"Ders '{row['Ders Adı']}' ({row['Ders Kodu']}) zaten var, işlem atlandı.")
            continue

        # Dersler tablosuna ekleme
        cursor.execute("""
            INSERT INTO Dersler (fakulte_id, bolum_id, sinif, ders_kodu, ders_adi, ogretim_uyesi_id, haftalik_saat, online, zorunlu_saat, statu)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (fakulte_id, bolum_id, row['Sınıf'], row['Ders Kodu'], row['Ders Adı'],
              ogretim_uyesi_id, row['Haftalık Saat'], online, row['Zorunlu Saat'], row['Statü']))

    conn.commit()
    cursor.close()
    conn.close()
    print("Ders ekleme işlemleri tamamlandı.")

insert_courses_from_excel(courses_file)  # Dersleri eklemek için


students_file = "Ogrenci_Ders.xlsx"
def insert_students_from_excel(students_file):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    df = pd.read_excel(students_file)

    for index, row in df.iterrows():
        numara = row['Numara']
        ders_adi = row['Ders Adı']

        # Ogrenci tablosundan ogrenci_id ve ogrenci_num al
        cursor.execute("SELECT id, numara FROM Ogrenciler WHERE numara = ?", (numara,))
        ogrenci = cursor.fetchone()

        if ogrenci:
            ogrenci_id = ogrenci[0]
            ogrenci_num = ogrenci[1]
        else:
            print(f"Hata: {numara} numaralı öğrenci bulunamadı.")
            continue

            # Dersler tablosundan ders_id al
        cursor.execute("SELECT id FROM Dersler WHERE ders_adi = ?", (ders_adi,))
        ders = cursor.fetchone()

        if ders:
            ders_id = ders[0]
        else:
            print(f"Hata: {ders_adi} adlı ders bulunamadı.")
            continue

            # Eğer öğrenci-ders ilişkisi zaten varsa işlemi atla
        cursor.execute("""
            SELECT id FROM OgrenciDers 
            WHERE ogrenci_id = ? AND ogrenci_num = ? AND ders_id = ?
        """, (ogrenci_id, ogrenci_num, ders_id))
        existing_record = cursor.fetchone()

        if existing_record:
            continue

        # OgrenciDers tablosuna ekleme yap
        cursor.execute("""
            INSERT INTO OgrenciDers (ogrenci_id, ogrenci_num, ders_id) 
            VALUES (?, ?, ?)
        """, (ogrenci_id, ogrenci_num, ders_id))

    conn.commit()
    cursor.close()
    conn.close()
    print("OgrenciDers ekleme işlemleri tamamlandı.")

insert_students_from_excel(students_file)  # Öğrencileri eklemek için


def add_faculty():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    fakulte_adi = input("Eklemek istediğiniz fakültenin adını girin: ").strip()

    # Fakülte zaten var mı kontrol et
    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    existing_faculty = cursor.fetchone()

    if existing_faculty:
        print(f"Hata: '{fakulte_adi}' fakültesi zaten mevcut.")
    else:
        cursor.execute("INSERT INTO Fakulte (fakulte_adi) VALUES (?)", (fakulte_adi,))
        conn.commit()
        print(f"'{fakulte_adi}' fakültesi başarıyla eklendi!")

    conn.close()


def delete_faculty():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    fakulte_adi = input("Silmek istediğiniz fakültenin adını girin: ").strip()

    # Fakülte ID'yi al
    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    faculty = cursor.fetchone()

    if not faculty:
        print(f"Hata: '{fakulte_adi}' fakültesi bulunamadı.")
        conn.close()
        return

    faculty_id = faculty[0]
    cursor.execute("SELECT COUNT(*) FROM Bolumler WHERE fakulte_id = ?", (faculty_id,))
    department_count = cursor.fetchone()[0]

    if department_count > 0:
        print(f"Hata: '{fakulte_adi}' fakültesine bağlı {department_count} bölüm bulunmaktadır. "
              "Önce bu bölümleri silmelisiniz.")
    else:
        cursor.execute("DELETE FROM Fakulte WHERE id = ?", (faculty_id,))
        conn.commit()
        print(f"'{fakulte_adi}' fakültesi başarıyla silindi.")

    conn.close()


def add_department():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    fakulte_adi = input("Bölüm eklemek istediğiniz fakültenin adını girin: ").strip()
    bolum_adi = input("Eklemek istediğiniz bölüm adını girin: ").strip()

    # Fakülte ID'yi al
    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    faculty = cursor.fetchone()

    if not faculty:
        print(f"Hata: '{fakulte_adi}' fakültesi bulunamadı. Önce fakülte ekleyin.")
        conn.close()
        return

    fakulte_id = faculty[0]
    # Bölümün zaten var olup olmadığını kontrol et
    cursor.execute("SELECT id FROM Bolumler WHERE bolum_adi = ? AND fakulte_id = ?", (bolum_adi, fakulte_id))
    existing_department = cursor.fetchone()

    if existing_department:
        print(f"Hata: '{bolum_adi}' bölümü zaten '{fakulte_adi}' fakültesinde mevcut.")
    else:
        cursor.execute(
            "INSERT INTO Bolumler (bolum_adi, fakulte_id, fakulte_adi) VALUES (?, ?, ?)",
            (bolum_adi, fakulte_id, fakulte_adi)
        )
        conn.commit()
        print(f"'{bolum_adi}' bölümü '{fakulte_adi}' fakültesine başarıyla eklendi!")

    conn.close()


def delete_department():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    fakulte_adi = input("Silmek istediğiniz fakültenin adını girin: ").strip()
    bolum_adi = input("Silmek istediğiniz bölüm adını girin: ").strip()

    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    faculty = cursor.fetchone()

    if not faculty:
        print(f"Hata: '{fakulte_adi}' fakültesi bulunamadı. Lütfen doğru fakülte adını girin.")
        conn.close()
        return

    # Silinecek bölümü kontrol et
    fakulte_id = faculty[0]
    cursor.execute("SELECT id FROM Bolumler WHERE bolum_adi = ? AND fakulte_id = ?", (bolum_adi, fakulte_id))
    department = cursor.fetchone()

    if not department:
        print(f"Hata: '{bolum_adi}' bölümü '{fakulte_adi}' fakültesinde bulunamadı.")
    else:
        cursor.execute("DELETE FROM Bolumler WHERE id = ?", (department[0],))
        conn.commit()
        print(f"'{bolum_adi}' bölümü '{fakulte_adi}' fakültesinden başarıyla silindi!")

    conn.close()


def add_instructor():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    ogretim_gorevlisi = input("Eklemek istediğiniz öğretim görevlisinin adını girin: ").strip()
    fakulte_adi = input("Öğretim görevlisinin bağlı olduğu fakülte adını girin: ").strip()

    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    faculty = cursor.fetchone()

    if not faculty:
        print(f"Hata: '{fakulte_adi}' fakültesi bulunamadı. Lütfen doğru fakülte adını girin.")
        conn.close()
        return

    fakulte_id = faculty[0]
    # Haftalık uygunluk saatlerini al
    pazartesi = input("Pazartesi uygunluk saatleri (örn: 10:00, 12:00, yoksa boş bırakın): ").strip() or None
    sali = input("Salı uygunluk saatleri (örn: 14:00, 16:00, yoksa boş bırakın): ").strip() or None
    carsamba = input("Çarşamba uygunluk saatleri (örn: 09:00, 11:00, yoksa boş bırakın): ").strip() or None
    persembe = input("Perşembe uygunluk saatleri (örn: 13:00, 15:00, yoksa boş bırakın): ").strip() or None
    cuma = input("Cuma uygunluk saatleri (örn: 15:00-17:00, yoksa boş bırakın): ").strip() or None

    # Öğretim görevlisini ekle
    try:
        cursor.execute("""
            INSERT INTO OgretimGorevlileri (ogretim_gorevlisi, fakulte_id, pazartesi, sali, carsamba, persembe, cuma)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (ogretim_gorevlisi, fakulte_id, pazartesi, sali, carsamba, persembe, cuma))

        conn.commit()
        print(f"'{ogretim_gorevlisi}', '{fakulte_adi}' fakültesine başarıyla eklendi!")

    except pyodbc.IntegrityError:
        print(f"Hata: '{ogretim_gorevlisi}' zaten mevcut. Lütfen farklı bir isim girin.")

    conn.close()


def delete_instructor():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    ogretim_gorevlisi = input("Silmek istediğiniz öğretim görevlisinin adını giriniz: ").strip()

    # Öğretim görevlisini arama
    cursor.execute("SELECT id FROM OgretimGorevlileri WHERE ogretim_gorevlisi = ?", (ogretim_gorevlisi,))
    instructor = cursor.fetchone()

    if not instructor:
        print(f"Hata: '{ogretim_gorevlisi}' öğretim görevlisi bulunamadı. Lütfen doğru isim girin.")
        conn.close()
        return

    # Öğretim görevlisini silme
    ogretim_uyesi_id = instructor[0]
    try:
        cursor.execute("DELETE FROM OgretimGorevlileri WHERE id = ?", (ogretim_uyesi_id,))
        conn.commit()
        print(f"'{ogretim_gorevlisi}' öğretim görevlisi başarıyla silindi!")

    except Exception as e:
        print(f"Bir hata oluştu: {e}")

    conn.close()


def add_student():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    numara = input("Öğrencinin numarasını girin: ").strip()
    fakulte_adi = input("Öğrencinin bağlı olduğu fakülte adını girin: ").strip()
    bolum_adi = input("Öğrencinin bağlı olduğu bölüm adını girin: ").strip()
    sinif = input("Öğrencinin sınıfını girin: ").strip()

    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    faculty = cursor.fetchone()

    if not faculty:
        print(f"Hata: '{fakulte_adi}' fakültesi bulunamadı. Lütfen doğru fakülte adını girin.")
        conn.close()
        return

    fakulte_id = faculty[0]
    # Bölüm ID'yi al
    cursor.execute("SELECT id FROM Bolumler WHERE bolum_adi = ? AND fakulte_id = ?", (bolum_adi, fakulte_id))
    department = cursor.fetchone()

    if not department:
        print(f"Hata: '{bolum_adi}' bölümü {fakulte_adi} fakültesinde bulunamadı. Lütfen doğru bölüm adını girin.")
        conn.close()
        return

    bolum_id = department[0]
    # Öğrenciyi ekle
    try:
        cursor.execute("""
            INSERT INTO Ogrenciler (numara, fakulte_id, bolum_id, bolum_adi, sinif)
            VALUES (?, ?, ?, ?, ?)
        """, (numara, fakulte_id, bolum_id, bolum_adi, sinif))

        conn.commit()
        print(f"'{numara}' numaralı öğrenci başarıyla {fakulte_adi} fakültesi, {bolum_adi} bölümüne eklendi!")

    except pyodbc.IntegrityError:
        print(f"Hata: '{numara}' numaralı öğrenci zaten mevcut. Lütfen farklı bir numara girin.")

    conn.close()


def delete_student():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    numara = input("Silmek istediğiniz öğrencinin numarasını girin: ").strip()

    # Öğrenci kaydını sil
    try:
        cursor.execute("DELETE FROM Ogrenciler WHERE numara = ?", (numara,))
        rows_affected = cursor.rowcount

        if rows_affected > 0:
            conn.commit()
            print(f"'{numara}' numaralı öğrenci başarıyla silindi.")
        else:
            print(f"Hata: '{numara}' numaralı öğrenci bulunamadı.")

    except pyodbc.Error as e:
        print(f"Veritabanı hatası: {e}")

    conn.close()


def add_classroom():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    derslik_id = input("Derslik ID'sini girin: ").strip()
    kapasite = input("Dersliğin kapasitesini girin: ").strip()

    # Kapasiteyi tam sayıya çevir
    try:
        kapasite = int(kapasite)
    except ValueError:
        print("Hata: Kapasite sayısal bir değer olmalıdır.")
        conn.close()
        return

    statu = input("Derslik statüsünü girin (LAB veya NORMAL): ").strip().upper()
    if statu not in ['LAB', 'NORMAL']:
        print("Hata: Statü 'LAB' veya 'NORMAL' olmalıdır.")
        conn.close()
        return

    try:
        cursor.execute("""
            INSERT INTO Derslikler (derslik_id, kapasite, statu)
            VALUES (?, ?, ?)
        """, (derslik_id, kapasite, statu))

        conn.commit()
        print(f"'{derslik_id}' ID'li derslik başarıyla eklendi!")

    except pyodbc.IntegrityError:
        print(f"Hata: '{derslik_id}' ID'li derslik zaten mevcut. Lütfen farklı bir derslik ID'si girin.")

    conn.close()


def delete_classroom():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    derslik_id = input("Silmek istediğiniz derslik ID'sini girin: ").strip()

    # Derslik var mı kontrolü
    cursor.execute("SELECT id FROM Derslikler WHERE derslik_id = ?", (derslik_id,))
    classroom = cursor.fetchone()

    if not classroom:
        print(f"Hata: '{derslik_id}' ID'li derslik bulunamadı. Lütfen doğru bir derslik ID'si girin.")
        conn.close()
        return

    # Dersliği silme
    try:
        cursor.execute("DELETE FROM Derslikler WHERE derslik_id = ?", (derslik_id,))
        conn.commit()
        print(f"'{derslik_id}' ID'li derslik başarıyla silindi!")

    except pyodbc.Error as e:
        print(f"Hata: Derslik silinirken bir problem oluştu: {e}")

    conn.close()


def add_course():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    ders_kodu = input("Dersin kodunu girin: ").strip()
    ders_adi = input("Dersin adını girin: ").strip()
    sinif = int(input("Sınıf bilgisini girin: ").strip())
    haftalik_saat = int(input("Haftalık saat bilgisini girin: ").strip())

    # Online durumunu al ve dönüştür
    online_input = input("Ders online mı? (Evet/Hayır): ").strip().casefold()
    if online_input in ["evet", "e"]:
        online = 1
    elif online_input in ["hayır", "h", "hayir"]:
        online = 0
    else:
        print("Hata: Lütfen 'Evet' veya 'Hayır' olarak cevap verin.")
        conn.close()
        return

    zorunlu_saat = input("Zorunlu saat bilgilerini girin (boş bırakabilirsiniz): ").strip()
    if not zorunlu_saat:
        zorunlu_saat = 0
    else:
        zorunlu_saat = int(zorunlu_saat)

    statu = input("Ders statüsünü girin (LAB/NORMAL): ").strip().upper()

    # Fakülte, Bölüm ve Öğretim Üyesi ID'lerini al
    fakulte_adi = input("Dersin bağlı olduğu fakülte adını girin: ").strip()
    bolum_adi = input("Dersin bağlı olduğu bölüm adını girin: ").strip()
    ogretim_uyesi_adi = input("Dersin öğretim üyesinin adını girin: ").strip()

    cursor.execute("SELECT id FROM Fakulte WHERE fakulte_adi = ?", (fakulte_adi,))
    fakulte = cursor.fetchone()

    if not fakulte:
        print(f"Hata: '{fakulte_adi}' fakültesi bulunamadı. Lütfen doğru fakülte adını girin.")
        conn.close()
        return

    fakulte_id = fakulte[0]
    cursor.execute("SELECT id FROM Bolumler WHERE bolum_adi = ?", (bolum_adi,))
    bolum = cursor.fetchone()

    if not bolum:
        print(f"Hata: '{bolum_adi}' bölümü bulunamadı. Lütfen doğru bölüm adını girin.")
        conn.close()
        return

    bolum_id = bolum[0]
    cursor.execute("SELECT id FROM OgretimGorevlileri WHERE ogretim_gorevlisi = ?", (ogretim_uyesi_adi,))
    ogretim_uyesi = cursor.fetchone()

    if not ogretim_uyesi:
        print(f"Hata: '{ogretim_uyesi_adi}' öğretim üyesi bulunamadı. Lütfen doğru öğretim üyesi adını girin.")
        conn.close()
        return

    ogretim_uyesi_id = ogretim_uyesi[0]
    # Ders ekle
    try:
        cursor.execute("""
            INSERT INTO Dersler (fakulte_id, bolum_id, sinif, ders_kodu, ders_adi, ogretim_uyesi_id, haftalik_saat, online, zorunlu_saat, statu)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
        fakulte_id, bolum_id, sinif, ders_kodu, ders_adi, ogretim_uyesi_id, haftalik_saat, online, zorunlu_saat, statu))

        conn.commit()
        print(f"'{ders_adi}' dersi başarıyla eklendi!")

    except pyodbc.Error as e:
        print(f"Hata: Ders eklenirken bir problem oluştu: {e}")

    conn.close()


def delete_course():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    ders_kodu = input("Silmek istediğiniz dersin kodunu girin: ").strip()

    # Ders kodunu kullanarak derse ait kaydı sil
    try:
        cursor.execute("SELECT * FROM Dersler WHERE ders_kodu = ?", (ders_kodu,))
        ders = cursor.fetchone()

        if not ders:
            print(f"Hata: '{ders_kodu}' dersine ait kayıt bulunamadı. Lütfen doğru ders kodu girin.")
            conn.close()
            return

        # Ders kaydını sil
        cursor.execute("DELETE FROM Dersler WHERE ders_kodu = ?", (ders_kodu,))
        conn.commit()
        print(f"'{ders_kodu}' ders kaydı başarıyla silindi!")

    except pyodbc.Error as e:
        print(f"Hata: Ders silinirken bir problem oluştu: {e}")

    conn.close()


def add_student_course():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    ogrenci_num = input("Öğrencinin numarasını girin: ").strip()
    ders_kodu = input("Ders kodunu girin: ").strip()

    cursor.execute("SELECT id FROM Ogrenciler WHERE numara = ?", (ogrenci_num,))
    ogrenci = cursor.fetchone()

    if not ogrenci:
        print(f"Hata: '{ogrenci_num}' numaralı öğrenci bulunamadı. Lütfen doğru öğrenci numarası girin.")
        conn.close()
        return

    ogrenci_id = ogrenci[0]
    # Ders koduna ait ders bilgilerini al
    cursor.execute("SELECT id FROM Dersler WHERE ders_kodu = ?", (ders_kodu,))
    ders = cursor.fetchone()

    if not ders:
        print(f"Hata: '{ders_kodu}' dersine ait kayıt bulunamadı. Lütfen doğru ders kodu girin.")
        conn.close()
        return

    ders_id = ders[0]
    # Öğrenci-ders ilişkisini ekle
    try:
        cursor.execute("""
            INSERT INTO OgrenciDers (ogrenci_id, ogrenci_num, ders_id)
            VALUES (?, ?, ?)
        """, (ogrenci_id, ogrenci_num, ders_id))

        conn.commit()
        print(f"Öğrenci '{ogrenci_num}' ile ders '{ders_kodu}' başarıyla ilişkilendirildi!")

    except pyodbc.IntegrityError as e:
        print(f"Hata: Öğrenci-ders ilişkisi eklenirken bir problem oluştu: {e}")

    conn.close()


def delete_student_course():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    ogrenci_num = input("Öğrencinin numarasını girin: ").strip()
    ders_kodu = input("Ders kodunu girin: ").strip()

    # Öğrenci numarasına ait öğrenci bilgilerini al
    cursor.execute("SELECT id FROM Ogrenciler WHERE numara = ?", (ogrenci_num,))
    ogrenci = cursor.fetchone()

    if not ogrenci:
        print(f"Hata: '{ogrenci_num}' numaralı öğrenci bulunamadı. Lütfen doğru öğrenci numarası girin.")
        conn.close()
        return

    ogrenci_id = ogrenci[0]
    # Ders koduna ait ders bilgilerini al
    cursor.execute("SELECT id FROM Dersler WHERE ders_kodu = ?", (ders_kodu,))
    ders = cursor.fetchone()

    if not ders:
        print(f"Hata: '{ders_kodu}' dersine ait kayıt bulunamadı. Lütfen doğru ders kodu girin.")
        conn.close()
        return

    ders_id = ders[0]
    # Öğrenci-ders ilişkisinin var olup olmadığını kontrol et
    cursor.execute("""
        SELECT * FROM OgrenciDers 
        WHERE ogrenci_id = ? AND ders_id = ?
    """, (ogrenci_id, ders_id))

    relation = cursor.fetchone()
    if not relation:
        print(f"Bu öğrenci ve ders arasında zaten bir ilişki bulunmamaktadır.")
        conn.close()
        return

    # İlişkiyi sil
    try:
        cursor.execute("""
            DELETE FROM OgrenciDers 
            WHERE ogrenci_id = ? AND ders_id = ?
        """, (ogrenci_id, ders_id))

        conn.commit()
        print(f"Öğrenci '{ogrenci_num}' ve ders '{ders_kodu}' arasındaki ilişki başarıyla silindi!")

    except pyodbc.Error as e:
        print(f"Hata: Öğrenci-ders ilişkisi silinirken bir problem oluştu: {e}")

    conn.close()


# Excel oluşturma işlemleri
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ders Programı"

days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
time_slots = ["09:00-10:00", "10:00-11:00", "11:00-12:00",
              "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
              "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"]
bm_class_headers = ["1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"]

# Hücre genişliklerini ayarlama
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15
for col in range(3, 3 + len(bm_class_headers)):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"))

# Başlıklar
ws.merge_cells("A1:B1")
ws["A1"] = "Bölüm"
ws["A1"].font = Font(bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:B2")
ws["A2"] = "Gün/Saatler"
ws["A2"].font = Font(bold=True)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("C1:F1")
ws["C1"] = "Bilgisayar Mühendisliği"
ws["C1"].font = Font(bold=True)
ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
ws["C1"].border = thin_border
ws["F1"].border = thin_border

# Sınıf başlıklarını ekleme
for idx, header in enumerate(bm_class_headers, start=3):
    cell = ws.cell(row=2, column=idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Gün ve saatleri ekleme
row_num = 3
for day in days:
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(time_slots) - 1, end_column=1)
    cell = ws.cell(row=row_num, column=1, value=day)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)
    cell.border = thin_border

    for time_slot in time_slots:
        cell = ws.cell(row=row_num, column=2, value=time_slot)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border
        row_num += 1

colors = ["FFDDC1", "D3E3FC", "FAF4B7", "D4E2D4"]
for row in ws.iter_rows(min_row=3, max_row=row_num - 1, min_col=3, max_col=6):
    for idx, cell in enumerate(row):
        cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
        cell.border = thin_border

row_num += 2
sw_class_headers = ["1. Sınıf", "2. Sınıf", "3. Sınıf"]
ws.merge_cells(f"A{row_num}:B{row_num}")
ws[f"A{row_num}"] = "Bölüm"
ws[f"A{row_num}"].font = Font(bold=True)
ws[f"A{row_num}"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(f"A{row_num + 1}:B{row_num + 1}")
ws[f"A{row_num + 1}"] = "Gün/Saatler"
ws[f"A{row_num + 1}"].font = Font(bold=True)
ws[f"A{row_num + 1}"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(f"C{row_num}:E{row_num}")
ws[f"C{row_num}"] = "Yazılım Mühendisliği"
ws[f"C{row_num}"].font = Font(bold=True)
ws[f"C{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
ws[f"C{row_num}"].border = thin_border

top_border = Border(top=Side(style="thin"), right=Side(style="thin"))
for col in range(1, 6):
    ws.cell(row=60, column=col).border = top_border

# Sınıf başlıklarını ekleme
for idx, header in enumerate(sw_class_headers, start=3):
    cell = ws.cell(row=row_num + 1, column=idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Gün ve saatleri ekleme
row_num += 2
for day in days:
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(time_slots) - 1, end_column=1)
    cell = ws.cell(row=row_num, column=1, value=day)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)
    cell.border = thin_border

    for time_slot in time_slots:
        cell = ws.cell(row=row_num, column=2, value=time_slot)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.border = thin_border
        row_num += 1

# Yazılım Mühendisliği için hücreleri renklendirme
for row in ws.iter_rows(min_row=row_num - len(days) * len(time_slots), max_row=row_num - 1, min_col=3, max_col=5):
    for idx, cell in enumerate(row):
        cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
        cell.border = thin_border

# Excel dosyasını kaydet
wb.save("Ders_Programi.xlsx")
time_slots = ["09:00-10:00", "10:00-11:00", "11:00-12:00",
              "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
              "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"]


# Online dersler
def get_online_courses():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    online_courses = []

    # Yanlızca online dersleri çekme işlevi
    cursor.execute("""
        SELECT ders_adi, haftalik_saat, ogretim_uyesi_id, sinif, bolum_id, zorunlu_saat
        FROM Dersler
        WHERE online = 1  -- SADECE ONLINE DERSLERİ AL
    """)

    online_courses = cursor.fetchall()
    conn.close()
    return online_courses


# Ortak dersleri belirleyip liste olarak döndür
def get_common_courses():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    common_courses = set()

    cursor.execute("""
        SELECT DISTINCT ders_adi, haftalik_saat, ogretim_uyesi_id, sinif
        FROM Dersler
        WHERE online = 0  -- Online dersleri hariç tut
        AND ders_adi IN (
            SELECT ders_adi FROM Dersler WHERE bolum_id IN (1, 2)
            GROUP BY ders_adi HAVING COUNT(DISTINCT bolum_id) = 2
        )
    """)

    for course in cursor.fetchall():
        course_name, hours_per_week, instructor_id, class_year = course
        common_courses.add((course_name, hours_per_week, instructor_id, class_year))

    conn.close()
    return list(common_courses)

#Belirtilen başlangıç ve bitiş saatleri arasındaki tüm saat aralıklarını oluşturur.
def expand_time_range(start_time, end_time):
    start_time = start_time.strip()
    end_time = end_time.strip()

    try:
        start = datetime.strptime(start_time, "%H:%M")
        end = datetime.strptime(end_time, "%H:%M")
    except ValueError as e:
        print(f"Hata: Saat formatı hatalı! start_time='{start_time}', end_time='{end_time}'")
        return []  # Hatalı veriyi boş listeyle dön

    slots = []
    while start < end:
        next_hour = start + timedelta(hours=1)
        slots.append(f"{start.strftime('%H:%M')}-{next_hour.strftime('%H:%M')}")
        start = next_hour

    return slots


def get_instructor_availability():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    instructor_availability = {}

    # Öğretim üyelerinin uygun saatlerini çek
    cursor.execute("""
        SELECT id, ogretim_gorevlisi, pazartesi, sali, carsamba, persembe, cuma
        FROM OgretimGorevlileri
    """)

    rows = cursor.fetchall()
    if not rows:
        print(" Uyarı: Öğretim üyelerinin uygun saatleri veritabanından çekilemedi!")
        return None

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

    for row in rows:
        instructor_id = row[0]
        instructor_name = row[1]  # Öğretim üyesi adı
        instructor_availability[instructor_id] = {}

        for i, day in enumerate(days, start=2):
            available_hours = row[i]
            if available_hours:
                times = available_hours.split(", ")
                expanded_slots = []

                # Saat aralıklarını genişletme işlevi
                for j in range(len(times) - 1):
                    expanded_slots.extend(expand_time_range(times[j], times[j + 1]))

                instructor_availability[instructor_id][day] = expanded_slots
            else:
                instructor_availability[instructor_id][day] = []

    conn.close()
    return instructor_availability


def sort_instructors_by_availability(instructor_availability):
    sorted_availability = {}
    for instructor, days in instructor_availability.items():
        sorted_days = sorted(days.items(), key=lambda x: len(x[1]), reverse=True)
        sorted_availability[instructor] = {day: times for day, times in sorted_days}
    return sorted_availability


# Uygun saatleri biçimlendiren fonksiyon
def convert_times_to_slots(instructor_availability, time_slots):
    print("time_slots içeriği:", time_slots)

    converted_availability = {}
    for instructor, availability in instructor_availability.items():
        converted_availability[instructor] = {}

        for day, hours in availability.items():
            converted_availability[instructor][day] = []

            for i in range(len(hours) - 1):
                start_time = hours[i]
                end_time = hours[i + 1]

                # Saat aralığı uygunsa time_slots'a ekle
                for slot in time_slots:
                    slot_start, slot_end = slot.split('-')
                    if slot_start == start_time and slot_end == end_time:
                        converted_availability[instructor][day].append(slot)

    return converted_availability

#Online ve ortak dersler dışında kalan bölüme özel dersleri veritabanından çeker.
def get_department_courses():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    department_courses = []

    cursor.execute("""
        SELECT DISTINCT ders_adi, haftalik_saat, ogretim_uyesi_id, sinif, bolum_id
        FROM Dersler
        WHERE online = 0  -- Online dersleri hariç tut
        AND ders_adi NOT IN (
            SELECT DISTINCT ders_adi FROM Dersler WHERE bolum_id IN (1, 2)
            GROUP BY ders_adi HAVING COUNT(DISTINCT bolum_id) = 2
        )  -- Ortak dersleri hariç tut
    """)

    for course in cursor.fetchall():
        course_name, hours_per_week, instructor_id, class_year, department_id = course
        department_courses.append((course_name, hours_per_week, instructor_id, class_year, department_id))

    conn.close()
    return department_courses


# Veritabanından gelen 'zorunlu_saat' değerlerini uygun 'time_slots' formatına çevirir.
def convert_mandatory_time(mandatory_time):
    slot_mapping = {
        "09:00": "09:00-10:00", "10:00": "10:00-11:00", "11:00": "11:00-12:00",
        "12:00": "12:00-13:00", "13:00": "13:00-14:00", "14:00": "14:00-15:00",
        "15:00": "15:00-16:00", "16:00": "16:00-17:00",
        "17:00": "17:00-18:00", "18:00": "18:00-19:00",
        "19:00": "19:00-20:00", "20:00": "20:00-21:00"
    }

    converted_slots = []
    times = sorted(mandatory_time.split(", "))  # Saatleri sıralı hale getirir

    for i in range(len(times) - 1):
        start_time = times[i]
        end_time = times[i + 1]

        # Eğer saatler ardışık değilse aradaki tüm saatleri ekler
        while start_time in slot_mapping and start_time != end_time:
            converted_slots.append(slot_mapping[start_time])
            hours, minutes = map(int, start_time.split(":"))
            hours += 1
            start_time = f"{hours:02d}:00"  # Yeni saati oluştur

    # Son saati de ekle
    if times[-1] in slot_mapping and slot_mapping[times[-1]] not in converted_slots:
        converted_slots.append(slot_mapping[times[-1]])

    return converted_slots


def get_instructor_name(instructor_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT ogretim_gorevlisi FROM OgretimGorevlileri WHERE id = ?
    """, (instructor_id,))
    row = cursor.fetchone()
    conn.close()

    return row[0] if row else "Bilinmeyen Öğretim Üyesi"


def assign_courses_to_schedule(online_courses, time_slots):
    wb = openpyxl.load_workbook("Ders_Programi.xlsx")
    ws = wb.active

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    # Boş program tablosu oluştur
    schedule = {
        day: {slot: {"Yazılım Mühendisliği": {1: None, 2: None, 3: None},
                     "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None}}
              for slot in time_slots}
        for day in days
    }

    # Dersin kaç saat atandığını takip eden sözlük işlevi
    assigned_hours_per_course = {course[0]: 0 for course in online_courses}

    # Öncelikle en fazla saat gerektiren dersleri sıralayan işlev (Büyükten küçüğe)
    online_courses.sort(key=lambda x: x[1], reverse=True)

    for course in online_courses:
        course_name, hours_per_week, instructor_id, class_year, department_id, mandatory_time = course
        instructor_name = get_instructor_name(instructor_id)

        # Eğer ders zaten atanmış ve haftalık saat dolmuşsa atamayı geç
        if assigned_hours_per_course[course_name] >= hours_per_week:
            continue

        assigned_hours = assigned_hours_per_course[course_name]
        assigned_slots = []

        # Dersin ait olduğu bölüm belirlenir
        department = "Bilgisayar Mühendisliği" if department_id == 2 else "Yazılım Mühendisliği"

        # Eğer ders her iki bölümde de aynı sınıfta okutuluyorsa bunu işaretle
        is_shared = any(c[0] == course_name and c[3] == class_year and c[4] != department_id for c in online_courses)

        # Zorunlu saatleri uygun formata çevir
        valid_mandatory_slots = convert_mandatory_time(mandatory_time)

        # Öncelikle zorunlu saatleri yerleştir
        for selected_day in days:
            for slot in valid_mandatory_slots:
                if assigned_hours >= hours_per_week:
                    break

                if schedule[selected_day][slot][department].get(class_year) is None:
                    schedule[selected_day][slot][department][class_year] = f"{course_name}\n{instructor_name}"
                    assigned_hours += 1
                    assigned_hours_per_course[course_name] = assigned_hours
                    assigned_slots.append(f"{selected_day}, {slot}, {class_year}. sınıf - {department}")

                    if is_shared:
                        other_department = "Bilgisayar Mühendisliği" if department == "Yazılım Mühendisliği" else "Yazılım Mühendisliği"
                        if schedule[selected_day][slot][other_department].get(class_year) is None:
                            schedule[selected_day][slot][other_department][
                                class_year] = f"{course_name}\n{instructor_name}"

        # Eğer hala boş saatler varsa kalanları yerleştir
        for selected_day in days:
            for slot in time_slots:
                if assigned_hours >= hours_per_week:
                    break

                if schedule[selected_day][slot][department].get(class_year) is None:
                    schedule[selected_day][slot][department][class_year] = course_name
                    assigned_hours += 1
                    assigned_hours_per_course[course_name] = assigned_hours
                    assigned_slots.append(f"{selected_day}, {slot}, {class_year}. sınıf - {department}")

                    if is_shared:
                        other_department = "Bilgisayar Mühendisliği" if department == "Yazılım Mühendisliği" else "Yazılım Mühendisliği"
                        if schedule[selected_day][slot][other_department].get(class_year) is None:
                            schedule[selected_day][slot][other_department][
                                class_year] = f"{course_name}\n{instructor_name}(Online)"

    # 4. Sınıfa 3. Sınıfın Derslerini Kopyala
    for day, slots in schedule.items():
        for slot, departments in slots.items():
            if departments["Bilgisayar Mühendisliği"][4] is None and departments["Bilgisayar Mühendisliği"][
                3] is not None:
                departments["Bilgisayar Mühendisliği"][4] = departments["Bilgisayar Mühendisliği"][3]

    # Debug için terminalde schedule'ı yazdır
    '''print("\n Debug İçin: Atama Sonrası Schedule Kontrolü\n")
    for day, slots in schedule.items():
        print(f" {day}:")
        for time_slot, classes in slots.items():
            for department, class_data in classes.items():
                for class_year, course in class_data.items():
                    if course:
                        print(f"   {time_slot} | {class_year}. sınıf - {department} -> {course}")'''

    wb.save("Ders_Programi.xlsx")

    # Dersleri excel'e yaz
    row_offsets = {
        "Pazartesi": 3,
        "Salı": 15,
        "Çarşamba": 27,
        "Perşembe": 39,
        "Cuma": 51
    }

    sw_row_offsets = {
        "Pazartesi": 67,
        "Salı": 79,
        "Çarşamba": 91,
        "Perşembe": 103,
        "Cuma": 115
    }

    class_columns_bm = {1: 3, 2: 4, 3: 5, 4: 6}  # Bilgisayar Mühendisliği için sınıf sütunları
    class_columns_sw = {1: 3, 2: 4, 3: 5}  # Yazılım Mühendisliği için sınıf sütunları

    for day, slots in schedule.items():
        for slot, classes in slots.items():
            try:
                slot_index = time_slots.index(slot)
            except ValueError:
                continue  # Eğer zaman dilimi bulunamazsa atla

            # Bilgisayar Mühendisliği'ni yaz
            row_num_bm = row_offsets[day] + slot_index
            for class_year, course_name in classes["Bilgisayar Mühendisliği"].items():
                if course_name and class_year in class_columns_bm:
                    col = class_columns_bm[class_year]
                    ws.cell(row=row_num_bm, column=col, value=f"{course_name}\n(Online)")
                    ws.cell(row=row_num_bm, column=col).alignment = Alignment(wrapText=True)

            # Yazılım Mühendisliği'ni yaz
            row_num_sw = sw_row_offsets[day] + slot_index
            for class_year, course_name in classes["Yazılım Mühendisliği"].items():
                if course_name and class_year in class_columns_sw:
                    col = class_columns_sw[class_year]
                    ws.cell(row=row_num_sw, column=col, value=f"{course_name}\n(Online)")
                    ws.cell(row=row_num_sw, column=col).alignment = Alignment(wrapText=True)

    wb.save("Ders_Programi.xlsx")


def assign_common_courses(common_courses, instructor_availability, time_slots):
    wb = openpyxl.load_workbook("Ders_Programi.xlsx")
    ws = wb.active
    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

    # Mevcut ders programını çek
    schedule = {
        day: {slot: {"Yazılım Mühendisliği": {1: None, 2: None, 3: None},
                     "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None}}
              for slot in time_slots}
        for day in days
    }

    # Mevcut excelden dersleri al ve schedule'a kaydet
    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col:
                    existing_bm_course = ws.cell(row=row_bm, column=bm_col).value
                    if existing_bm_course:
                        schedule[day][slot]["Bilgisayar Mühendisliği"][class_year] = existing_bm_course

                if sw_col:
                    existing_sw_course = ws.cell(row=row_sw, column=sw_col).value
                    if existing_sw_course:
                        schedule[day][slot]["Yazılım Mühendisliği"][class_year] = existing_sw_course

    # Öğretim üyesinin uygun saatlerinin belirlenmesi
    instructor_schedule = {instructor: {day: [] for day in days} for instructor in instructor_availability}

    # Sözlük ile derslerin kaç saat atandığını kontrol etme
    assigned_hours_per_course = {course[0]: 0 for course in common_courses}

    # Ortak dersleri öncelik sırasına göre ekleme
    common_courses.sort(key=lambda x: (x[3], -x[1]))

    for course in common_courses:
        course_name, hours_per_week, instructor_id, class_year = course
        instructor_name = get_instructor_name(instructor_id)

        if instructor_id not in instructor_availability:
            print(f" Öğretim Üyesi ID {instructor_id} için uygunluk bilgisi bulunamadı. {course_name} atlanıyor.")
            continue

        if assigned_hours_per_course[course_name] >= hours_per_week:
            continue

        assigned_hours = 0
        assigned_slots = []

        # Ortak dersi alan sınıfların bulunması
        related_classes = [c[3] for c in common_courses if c[0] == course_name]

        best_block = None
        for selected_day in days:
            if selected_day not in instructor_availability[instructor_id]:
                continue

            available_slots = instructor_availability[instructor_id][selected_day]
            # Boş olan saaatlerin bulunması
            for i in range(len(available_slots) - (hours_per_week - 1)):
                block_slots = available_slots[i:i + hours_per_week]

                is_valid = all([
                    slot in time_slots for slot in block_slots
                ]) and all([
                    slot not in instructor_schedule[instructor_id][selected_day] for slot in block_slots
                ]) and all([
                    all(schedule[selected_day][slot]["Bilgisayar Mühendisliği"].get(cls) is None and
                        schedule[selected_day][slot]["Yazılım Mühendisliği"].get(cls) is None
                        for cls in related_classes)
                    for slot in block_slots
                ])

                if is_valid:
                    best_block = (selected_day, block_slots)
                    break  # Uygun ilk blok bulunduğunda döngüden çıkılır

        # Eğer en iyi blok bulunduysa ders atanır
        if best_block:
            selected_day, block_slots = best_block

            for slot in block_slots:
                for cls in related_classes:
                    schedule[selected_day][slot]["Bilgisayar Mühendisliği"][cls] = f"{course_name}\n{instructor_name}"
                    schedule[selected_day][slot]["Yazılım Mühendisliği"][cls] = f"{course_name}\n{instructor_name}"

                instructor_schedule[instructor_id][selected_day].append(slot)
                assigned_hours += 1
                assigned_hours_per_course[course_name] = assigned_hours
                assigned_slots.append(f"{selected_day}, {slot}, {related_classes} sınıfları - Ortak Ders")

    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col and schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]
                    ws.cell(row=row_bm, column=bm_col, value=f"{course_name}\n")
                    ws.cell(row=row_bm, column=bm_col).alignment = Alignment(wrapText=True)

                if sw_col and schedule[day][slot]["Yazılım Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Yazılım Mühendisliği"][class_year]
                    ws.cell(row=row_sw, column=sw_col, value=f"{course_name}\n")
                    ws.cell(row=row_sw, column=sw_col).alignment = Alignment(wrapText=True)

    wb.save("Ders_Programi.xlsx")


#Bölüme özel dersleri uygun boş saatlere yerleştirir ve excel'e kaydeder.
def assign_department_courses(department_courses, instructor_availability, time_slots):
    wb = openpyxl.load_workbook("Ders_Programi.xlsx")
    ws = wb.active

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

    # Mevcut ders programını çek (Online ve ortak derslerin olduğu program)
    schedule = {
        day: {slot: {"Yazılım Mühendisliği": {1: None, 2: None, 3: None},
                     "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None}}
              for slot in time_slots}
        for day in days
    }

    # Mevcut excelden dersleri al ve schedule'a kaydet
    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col:
                    existing_bm_course = ws.cell(row=row_bm, column=bm_col).value
                    if existing_bm_course:
                        schedule[day][slot]["Bilgisayar Mühendisliği"][class_year] = existing_bm_course

                if sw_col:
                    existing_sw_course = ws.cell(row=row_sw, column=sw_col).value
                    if existing_sw_course:
                        schedule[day][slot]["Yazılım Mühendisliği"][class_year] = existing_sw_course

    # Öğretim üyelerinin uygun saatlerini alır
    instructor_schedule = {instructor: {day: [] for day in days} for instructor in instructor_availability}

    # Belirtilen öğretim üyesinin bu saat diliminde başka sınıfta dersi olup olmadığını kontrol eder.
    def is_instructor_available(instructor_name, selected_day, slot):
        for other_department in ["Bilgisayar Mühendisliği", "Yazılım Mühendisliği"]:
            for other_class in range(1, 5):
                assigned_course = schedule[selected_day][slot][other_department].get(other_class)
                if assigned_course and instructor_name in assigned_course:  
                    return False  # Eğitmen bu saatte uygun değil
        return True  # Eğitmen bu saatte uygun

    # Bölüm derslerini öncelik sırasına göre sırala (Saat sayısına göre büyükten küçüğe)
    department_courses.sort(key=lambda x: x[1], reverse=True)

    for course in department_courses:
        course_name, hours_per_week, instructor_id, class_year, department_id = course
        instructor_name = get_instructor_name(instructor_id)

        if instructor_id not in instructor_availability:
            print(f"Öğretim Üyesi ID {instructor_id} için uygunluk bilgisi bulunamadı. {course_name} atlanıyor.")
            continue

        assigned_hours = 0
        assigned_slots = []
        department = "Bilgisayar Mühendisliği" if department_id == 2 else "Yazılım Mühendisliği"

        best_block = None
        for selected_day in days:
            available_slots = instructor_availability.get(instructor_id, {}).get(selected_day, [])

            for i in range(len(available_slots) - (hours_per_week - 1)):
                block_slots = available_slots[i:i + hours_per_week]
                is_valid = all([
                    slot in time_slots for slot in block_slots
                ]) and all([
                    is_instructor_available(instructor_name, selected_day, slot) for slot in block_slots
                ]) and all([
                    schedule[selected_day][slot][department].get(class_year) is None for slot in block_slots
                ])

                if is_valid:
                    best_block = (selected_day, block_slots)
                    break

        if best_block:
            selected_day, block_slots = best_block
            for slot in block_slots:
                schedule[selected_day][slot][department][class_year] = f"{course_name}\n{instructor_name}"
                instructor_schedule[instructor_id][selected_day].append(slot)
                assigned_hours += 1
                assigned_slots.append(f"{selected_day}, {slot}, {class_year}. sınıf - {department}")

    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col and schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]
                    ws.cell(row=row_bm, column=bm_col, value=f"{course_name}\n")
                    ws.cell(row=row_bm, column=bm_col).alignment = Alignment(wrapText=True)

                if sw_col and schedule[day][slot]["Yazılım Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Yazılım Mühendisliği"][class_year]
                    ws.cell(row=row_sw, column=sw_col, value=f"{course_name}\n")
                    ws.cell(row=row_sw, column=sw_col).alignment = Alignment(wrapText=True)

    wb.save("Ders_Programi.xlsx")


# Derslikler
def get_classrooms():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT derslik_id, kapasite, statu FROM Derslikler
    """)

    classrooms = []
    for row in cursor.fetchall():
        classrooms.append({
            "id": row[0],
            "capacity": row[1],
            "status": row[2]
        })

    conn.close()
    return classrooms


# Dersi alan öğrenci sayısı
def get_student_count_for_course(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*) FROM OgrenciDers WHERE ders_id = ?
    """, (course_id,))

    row = cursor.fetchone()
    conn.close()

    return row[0] if row else 0


# Dersler
def get_course_id(course_name):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id FROM Dersler WHERE ders_adi = ?
    """, (course_name,))

    row = cursor.fetchone()
    conn.close()

    return row[0] if row else None


# Online dersler
def get_online_status(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("SELECT online FROM Dersler WHERE id = ?", (course_id,))
    row = cursor.fetchone()
    conn.close()

    return int(row[0])


# Kayıtlı excel dosyası
def read_courses_from_excel(filename="Ders_Programi.xlsx"):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    schedule = {}

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    row_offsets = {"Pazartesi": 3, "Salı": 15, "Çarşamba": 27, "Perşembe": 39, "Cuma": 51}
    sw_row_offsets = {"Pazartesi": 67, "Salı": 79, "Çarşamba": 91, "Perşembe": 103, "Cuma": 115}

    class_columns_bm = {1: 3, 2: 4, 3: 5, 4: 6}
    class_columns_sw = {1: 3, 2: 4, 3: 5}

    for day in days:
        schedule[day] = {}
        for slot in time_slots:
            schedule[day][slot] = {
                "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None},
                "Yazılım Mühendisliği": {1: None, 2: None, 3: None}
            }

            slot_index = time_slots.index(slot)
            row_bm = row_offsets[day] + slot_index
            row_sw = sw_row_offsets[day] + slot_index

            for class_year in range(1, 5):
                bm_col = class_columns_bm.get(class_year, None)
                if bm_col:
                    cell_value = ws.cell(row=row_bm, column=bm_col).value
                    if cell_value:
                        schedule[day][slot]["Bilgisayar Mühendisliği"][class_year] = cell_value.strip()

            for class_year in range(1, 4):
                sw_col = class_columns_sw.get(class_year, None)
                if sw_col:
                    cell_value = ws.cell(row=row_sw, column=sw_col).value
                    if cell_value:
                        schedule[day][slot]["Yazılım Mühendisliği"][class_year] = cell_value.strip()

    return schedule


# Ders statülerini belirleme
def get_course_status(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT statu FROM Dersler WHERE id = ?
    """, (course_id,))

    row = cursor.fetchone()
    conn.close()

    return row[0] if row else "NORMAL"


# Haftalık toplam ders saati
def get_course_duration(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("SELECT haftalik_saat FROM Dersler WHERE id = ?", (course_id,))
    row = cursor.fetchone()
    conn.close()

    return int(row[0])


# Derslik ataması
def assign_classrooms_to_courses(schedule, time_slots):
    classrooms = get_classrooms()
    course_classroom_map = {}  # Her dersin ilk atanan dersliğini tutar
    occupied_classrooms = {}  # Hangi dersliklerin dolu olduğunu tutar
    course_duration_map = {}  # Hangi dersin kaç saat sürdüğünü tutar

    for day, slots in schedule.items():
        for slot_index, slot in enumerate(time_slots):
            for department, classes in schedule[day][slot].items():
                for class_year, course_info in classes.items():
                    if not course_info:
                        continue

                    lines = course_info.split("\n")
                    course_name = lines[0]
                    instructor_name = lines[1] if len(lines) > 1 else "Bilinmeyen Eğitmen"

                    # Dersin ID'si
                    course_id = get_course_id(course_name)
                    if not course_id:
                        print(f"⚠️ {course_name} için ders bulunamadı, ancak derslik ataması devam ediyor.")
                        continue

                    # Ders online mı kontrolü
                    online_status = get_online_status(course_id)
                    if int(online_status) == 1:
                        schedule[day][slot][department][class_year] = f"{course_name}\n{instructor_name} (Online)"
                        continue

                    # Dersin kaç saat olduğu bilgisi
                    if course_name not in course_duration_map:
                        course_duration_map[course_name] = get_course_duration(course_id)
                    duration = course_duration_map[course_name]

                    if course_name in course_classroom_map:
                        classroom_id = course_classroom_map[course_name]
                    else:
                        course_status = get_course_status(course_id)
                        student_count = get_student_count_for_course(course_id)  #Dersin öğrenci sayısı

                        # Ders için uygun kapasitedeki derslikleri filtreleme
                        suitable_classrooms = sorted(
                            [c for c in classrooms if c["capacity"] >= student_count and c["status"] == course_status],
                            key=lambda x: x["capacity"])

                        if not suitable_classrooms:
                            print(f"{course_name} için uygun derslik bulunamadı!")
                            continue

                        # İlk uygun ve müsait dersliği seç
                        classroom_id = None
                        for classroom in suitable_classrooms:
                            # Eğer bu derslik başka bir dersin saatleri içinde doluysa geç
                            is_available = True
                            for i in range(duration):  # Dersin süresi boyunca kontrol et
                                future_slot = time_slots[slot_index + i] if slot_index + i < len(time_slots) else None
                                if future_slot and (day, future_slot) in occupied_classrooms and classroom["id"] in \
                                        occupied_classrooms[(day, future_slot)]:
                                    is_available = False
                                    break

                            if is_available:
                                classroom_id = classroom["id"]
                                course_classroom_map[course_name] = classroom_id
                                break
                        else:
                            continue

                    # Dersliğin bu gün ve saatte dolu olduğunu kaydet
                    for i in range(duration):
                        future_slot = time_slots[slot_index + i] if slot_index + i < len(time_slots) else None
                        if future_slot:
                            if (day, future_slot) not in occupied_classrooms:
                                occupied_classrooms[(day, future_slot)] = set()
                            occupied_classrooms[(day, future_slot)].add(classroom_id)

                    # Programda dersi güncelle ve dersliği ekle
                    schedule[day][slot][department][class_year] = f"{course_name}\n{instructor_name} ({classroom_id})"

    return schedule


def main():
    print("\n📌 Dersler veritabanından çekiliyor.")
    online_courses = get_online_courses()   #Online dersler veritabanından çekiliyor
    common_courses = get_common_courses()  #Ortak dersler veritabanından çekiliyor
    department_courses = get_department_courses() #Bölüme özel dersler veritabanından çekiliyor

    print("📌 Öğretim üyelerinin uygunluk durumu alınıyor.")
    instructor_availability = get_instructor_availability()
    if instructor_availability is None:
        print("Hata: Öğretim üyesi uygunluk verisi çekilemedi. Veritabanını kontrol et!")
        exit(1)

    time_slots = [
        "09:00-10:00", "10:00-11:00", "11:00-12:00",
        "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
        "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"
    ]

    print("📌 Ders programı oluşturuluyor ve Excel'e kaydediliyor.")

    print("\n📌 Dersler Atanıyor.")
    # Online Dersleri Atama İşlevi
    if online_courses:
        assign_courses_to_schedule(online_courses, time_slots)
    else:
        print("Atanacak online ders bulunamadı!")

    # Ortak Dersleri Atama İşlevi
    if common_courses:
        assign_common_courses(common_courses, instructor_availability, time_slots)
    else:
        print("Atanacak ortak ders bulunamadı!")

    # Bölüme Özel Dersleri Atama İşlevi
    if department_courses:
        assign_department_courses(department_courses, instructor_availability, time_slots)
    else:
        print("Atanacak bölüme özel ders bulunamadı!")

    # Kaydedilen Excel'den Dersleri Okuma İşlevi
    schedule = read_courses_from_excel()

    # Derslik Ataması Yapma İşlevi
    print("📌 Derslikler Atanıyor.")
    schedule = assign_classrooms_to_courses(schedule, time_slots)

    # Güncellenmiş Programı Excel'e Kaydet
    print("\n✅ Ders Programı Excel'e Kaydediliyor.")
    wb = openpyxl.load_workbook("Ders_Programi.xlsx")
    ws = wb.active

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    row_offsets = {"Pazartesi": 3, "Salı": 15, "Çarşamba": 27, "Perşembe": 39, "Cuma": 51}
    sw_row_offsets = {"Pazartesi": 67, "Salı": 79, "Çarşamba": 91, "Perşembe": 103, "Cuma": 115}

    class_columns_bm = {1: 3, 2: 4, 3: 5, 4: 6}
    class_columns_sw = {1: 3, 2: 4, 3: 5}

    for day, slots in schedule.items():
        for slot, classes in slots.items():
            try:
                slot_index = time_slots.index(slot)
            except ValueError:
                continue

            row_num_bm = row_offsets[day] + slot_index
            row_num_sw = sw_row_offsets[day] + slot_index

            # Bilgisayar Mühendisliği'ni yaz
            for class_year, course_name in classes["Bilgisayar Mühendisliği"].items():
                if course_name and class_year in class_columns_bm:
                    ws.cell(row=row_num_bm, column=class_columns_bm[class_year], value=f"{course_name}\n")
                    ws.cell(row=row_num_bm, column=class_columns_bm[class_year]).alignment = Alignment(wrapText=True)

            # Yazılım Mühendisliği'ni yaz
            for class_year, course_name in classes["Yazılım Mühendisliği"].items():
                if course_name and class_year in class_columns_sw:
                    ws.cell(row=row_num_sw, column=class_columns_sw[class_year], value=f"{course_name}\n")
                    ws.cell(row=row_num_sw, column=class_columns_sw[class_year]).alignment = Alignment(wrapText=True)

    wb.save("Ders_Programi.xlsx")


def menu():
    while True:
        print("\nİşlemi Seçiniz")
        print("Fakülte İşlemleri:")
        print("1. Fakülte Ekle")
        print("2. Fakülte Sil")
        print("\nBölüm İşlemleri:")
        print("3. Bölüm Ekle")
        print("4. Bölüm Sil")
        print("\nÖğretim Üyesi İşlemleri:")
        print("5. Öğretim Görevlisi Ekle")
        print("6. Öğretim Görevlisi Sil")
        print("\nÖğrenci İşlemleri:")
        print("7. Öğrenci Ekle")
        print("8. Öğrenci Sil")
        print("\nDerslik İşlemleri:")
        print("9. Derslik Ekle")
        print("10. Derslik Sil")
        print("\nDers İşlemleri:")
        print("11. Ders Ekle")
        print("12. Ders Sil")
        print("\nÖğrenci-Ders İşlemleri:")
        print("13. Öğrenci-Ders Ekle")
        print("14. Öğrenci-Ders Sil")
        print("\n15. Çıkış")

        choice = input("Enter your choice (1-15): ").strip()

        if choice == '1':
            add_faculty()
        elif choice == '2':
            delete_faculty()
        elif choice == '3':
            add_department()
        elif choice == '4':
            delete_department()
        elif choice == '5':
            add_instructor()
        elif choice == '6':
            delete_instructor()
        elif choice == '7':
            add_student()
        elif choice == '8':
            delete_student()
        elif choice == '9':
            add_classroom()
        elif choice == '10':
            delete_classroom()
        elif choice == '11':
            add_course()
        elif choice == '12':
            delete_course()
        elif choice == '13':
            add_student_course()
        elif choice == '14':
            delete_student_course()
        elif choice == '15':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 15.")

menu()
main()